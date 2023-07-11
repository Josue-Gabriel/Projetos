import PySimpleGUI as sg
import csv
import locale
import sys
import openpyxl

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
sg.theme('DarkPurple')
caminho = "EXP_2022.csv"
regiao = ''
lista_geral = []
lista_das_regioes = []
ligado = True
pesquisas = False

valor = None
mes = None
CO_NCM = None
caminho = None
ativo = False
meses = {
    1: 'Janeiro',
    2: 'Fevereiro',
    3: 'Março',
    4: 'Abril',
    5: 'Maio',
    6: 'Junho',
    7: 'Julho',
    8: 'Agosto',
    9: 'Setembro',
    10: 'Outubro',
    11: 'Novembro',
    12: 'Dezembro'
}

def arquivo():
    global mes
    global CO_NCM
    global valor
    global caminho
    caminho_do_arquivo = [
        [sg.Text('Caminho de acesso ao arquivo', pad=(70,0), text_color='yellow', key='mensagem')],
        [sg.Input(key='Caminho', text_color='yellow')],
        [sg.Button('Buscar', key='Busca', pad=(140,0))]
    ]

    window = sg.Window('', caminho_do_arquivo)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Busca':
            try:
                caminho = values['Caminho']
                with open(caminho, 'r', newline='') as arquivo_csv:
                    leitor_csv = csv.reader(arquivo_csv, delimiter=';')
                    next(leitor_csv)
                    valores_do_arquivo = next(leitor_csv)
                mes = int(valores_do_arquivo[1])
                CO_NCM = valores_do_arquivo[2]
                valor = int(''.join(valores_do_arquivo[-1]))
                valor = valor / 100
                valor = locale.currency(valor, grouping=True, symbol=True)
                valor = str(valor)
                window.close()
            except:
                pass
            window['mensagem'].update('Arquivo não encontrado')
    try:
        mes = meses[mes]
        return mes, CO_NCM, valor
    except:
        sys.exit()

def tela_principal():
    global ativo
    global mes
    global CO_NCM
    global valor
    global pesquisas
    tela_principal = [
        [sg.Button('<', key='Retornar'),sg.Text('SP 2022', relief=sg.RELIEF_RIDGE, pad=(100,0), background_color=('#00FFFF'), text_color='black')],
        [sg.Text('Mês', size=(28, 0), text_color='black'), sg.Text(mes, key='mes', text_color='yellow', relief=sg.RELIEF_RIDGE)],
        [sg.Text('CO_NCM', size=(28, 0), text_color='black'), sg.Text(CO_NCM, key='CO_NCM', text_color='yellow', relief=sg.RELIEF_RIDGE)],
        [sg.Text('Valor', size=(28, 0), text_color='black'), sg.Button(valor[0:4] + '...', key='Valor')],
        [sg.Button('<', key='Voltar', size=(3, 1)), sg.Button('Pesquisa', key='Pesquisa',size=(26, 0)), sg.Button('>', key='Proxima', size=(3, 1))]
    ]

    janela = sg.Window(caminho, tela_principal)

    with open(caminho, 'r', newline='') as arquivo_csv:
        leitor_csv = csv.reader(arquivo_csv, delimiter=';')
        linhas = list(leitor_csv)
        num_linha = len(linhas)
        indice_atual = 1
        itens_separados = ', '.join(str(item) for item in linhas[indice_atual])
        lista_dos_valores = itens_separados.split(', ')
        completo = 0
        while True:
            event, values = janela.read()
            if event == sg.WIN_CLOSED:
                break
            if event == 'Retornar':
                janela.close()
                ativo = True
            if event == 'Proxima':
                indice_atual += 1
                if indice_atual >= num_linha:
                    indice_atual = num_linha - 1
                itens_separados = ', '.join(str(item) for item in linhas[indice_atual])
                lista_dos_valores = itens_separados.split(', ')
                mes = int(lista_dos_valores[1])
                CO_NCM = lista_dos_valores[2]
                janela['mes'].update(meses[mes])
                janela['CO_NCM'].update(CO_NCM)
                valor = int(''.join(lista_dos_valores[-1]))
                valor = valor / 100
                valor = locale.currency(valor, symbol=True, grouping=False, international=False)
                valor = str(valor)
                if completo == 0:
                    janela['Valor'].update(valor[0:4] + "...")
                if completo == 1:
                    janela['Valor'].update(valor)
                if completo == 2:
                    janela['Valor'].update(valor[0:4] + "...")
            if event == 'Voltar':
                indice_atual -= 1
                if indice_atual <= 0:
                    indice_atual = 1
                itens_separados = ', '.join(str(item) for item in linhas[indice_atual])
                lista_dos_valores = itens_separados.split(', ')
                mes = int(lista_dos_valores[1])
                CO_NCM = lista_dos_valores[2]
                janela['mes'].update(meses[mes])
                janela['CO_NCM'].update(CO_NCM)
                valor = int(''.join(lista_dos_valores[-1]))
                valor = valor / 100
                valor = locale.currency(valor, symbol=True, grouping=False, international=False)
                valor = str(valor)
                if completo == 0:
                    janela['Valor'].update(valor[0:4] + "...")
                if completo == 1:
                    janela['Valor'].update(valor)
                if completo == 2:
                    janela['Valor'].update(valor[0:4] + "...")
            if event == 'Valor':
                completo += 1
                if completo == 1:
                    janela['Valor'].update(valor)
                if completo == 2:
                    janela['Valor'].update(valor[0:4] + "...")
                    completo = 0
            if event == 'Pesquisa':
                ativo = False
                pesquisas = True
                janela.close()

arquivo()
tela_principal()

while ativo == True:
    arquivo()
    tela_principal()

if pesquisas == True:
    with open(caminho, 'r', newline='') as arquivo_csv:
        leitor_csv = csv.reader(arquivo_csv, delimiter=';')
        linhas = list(leitor_csv)
        num_linas = len(linhas)
        contador = 1
        while contador<num_linas:
            itens_separados = ', '.join(str(item) for item in linhas[contador])
            lista_dos_valores = itens_separados.split(', ')
            contador = contador + 1
            if lista_dos_valores[5] in lista_das_regioes:
                pass
            else:
                lista_das_regioes.append(lista_dos_valores[5])

    regioes = ', '.join(lista_das_regioes)

    def selecionar_regiao():
        global regiao
        global regioes
        global lista_das_regioes
        escolha_da_regiao = [
            [sg.Text("Selecione uma das opções de busca", relief=sg.RELIEF_RIDGE, key="texto")],
            [sg.Text(regioes, size=(40, 10))],
            [sg.Input(key="escolha", size=(30, 10)), sg.Button("Selecionar")]
        ]

        janela_regiao = sg.Window('', escolha_da_regiao)

        while True:
            acao, valor = janela_regiao.read()
            if acao == sg.WINDOW_CLOSED:
                break
            if acao == "Selecionar":
                regiao = valor['escolha']
                regiao = str(regiao)
                if regiao in lista_das_regioes:
                    janela_regiao.close()
                else:
                    janela_regiao['texto'].update("Por favor, selecione uma das opções abaixo")
        return regiao

    indice = 0
    lin = 1
    number = 1
    number_mes = number
    number = str(number)
    number = "0" + number
    itens_separados = ', '.join(str(item) for item in linhas[indice])
    lista_dos_valores = itens_separados.split(', ')

    regiao = selecionar_regiao()

    def busca():
        global indice
        global lin
        global itens_separados
        global lista_dos_valores
        global lista_geral
        global number
        try:
            while indice<num_linas:
                if number in lista_dos_valores[1]:
                    if regiao in lista_dos_valores[5]:
                        valor = ''.join(lista_dos_valores[-1])
                        valor = int(valor)
                        valor = valor / 100
                        valor_formatado = locale.currency(valor, grouping=True, symbol=True)
                        valor_final = str(valor_formatado)
                        lista_geral.append(lista_dos_valores[2] + " "*70 + valor_final)
                indice += 1
                lin += 1
                itens_separados = ', '.join(str(item) for item in linhas[indice])
                lista_dos_valores = itens_separados.split(', ')
        except:
            pass
        return lista_geral

    resultado = "\n".join(busca())

    def janela_pesquisa():
        global indice
        global lin
        global lista_geral
        global resultado
        global number
        global ativo
        global listagem
        layout = [
            [sg.Button("Voltar"), sg.Text(meses[number_mes], pad=(150,0), key="mes"), sg.Image(source='icone.png', key='planilha', enable_events=True)],
            [sg.Text("Nº Produto", size=(20, 0), key="numero"), sg.Text("", size=(24, 0), key="regiao"), sg.Text("Valor", key="novo_mes")],
            [sg.Multiline(resultado, size=(60, 10), key="listas")],
            [sg.Button("<", size=(2, 1), key="Retornar"), sg.Button(">", size=(2, 1), key="Proximo"), sg.Input(key="busca", size=(44,0)), sg.Button("Buscar")]
        ]

        janela = sg.Window("", layout)

        while True:
            acao, valor = janela.read()
            if acao == sg.WINDOW_CLOSED:
                sys.exit()
            if acao == "Proximo":
                indice = 0
                lin = 1
                number = int(number)
                number = number + 1
                if number <= 12:
                    janela["mes"].update(meses[number])
                    if number>=10:
                        number = str(number)
                    else:
                        number = str(number)
                        number = "0" + number
                    lista_geral = []
                    resultado = "\n".join(busca())
                    janela['listas'].update(resultado)
                else:
                    number = number - 1

            if acao == "Retornar":
                indice = 0
                lin = 1
                number = int(number)
                number = number - 1
                if number >= 1:
                    janela["mes"].update(meses[number])
                    if number>=10:
                        number = str(number)
                    else:
                        number = str(number)
                        number = "0" + number
                    lista_geral = []
                    resultado = "\n".join(busca())
                    janela['listas'].update(resultado)
                else:
                    number = number + 1
                
            if acao == "Buscar":
                buscas = valor['busca']
                buscas = str(buscas)
                contador = 0
                lista_geral = []
                while contador<num_linas:
                    itens_separados = ', '.join(str(item) for item in linhas[contador])
                    lista_dos_valores = itens_separados.split(', ')
                    if buscas in lista_dos_valores:
                        valor = ''.join(lista_dos_valores[-1])
                        valor = int(valor)
                        valor = valor / 100
                        number = lista_dos_valores[1]
                        number = int(number)
                        mes = meses[number]
                        lista_geral.append(lista_dos_valores[2] + " "*32 + lista_dos_valores[5] + " "*35 + mes)
                    else:
                        pass
                    contador += 1
                resultado = "\n".join(lista_geral)
                janela['listas'].update(resultado)
                janela['mes'].update("Resultado") 
                janela['regiao'].update("Região")
                janela['novo_mes'].update("mes")

            if acao == "Voltar":
                break
            
            if acao == "planilha":
                try:
                    val = ', '.join(linhas[1])
                    val = val.split(', ')

                    workbook = openpyxl.Workbook()

                    planilha = workbook.active
                    
                    lin = 1
                    limite = 1
                    contador = 1
                    
                    colunas_dos_meses = ['CO_NCM', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
                    
                    for i, valor in enumerate(colunas_dos_meses, start=1):
                        planilha.cell(row=1, column=contador, value=valor)
                        contador = contador + 1

                    contador = 1
                    linha = 2
                    while limite < num_linas:
                        if "SP" in val:
                            for i, valor in enumerate(val[2:3], start=1):
                                planilha.cell(row=linha, column=1, value=valor)
                                contador = contador + 1
                                for i, valor in enumerate(val[10:11], start=1):
                                    colum = int(val[1])
                                    colum = colum + 1
                                    valor_final = int(val[-1])
                                    valor_final = valor_final / 100
                                    valor_final = locale.currency(valor_final, symbol=True, grouping=False, international=False)
                                    valor_final = str(valor_final)
                                    planilha.cell(row=linha, column=colum, value=valor_final)
                            linha = linha + 1
                        val = ', '.join(linhas[lin])
                        val = val.split(', ')
                        lin = lin + 1
                        limite = limite + 1
                        contador = 1
                    
                    workbook.save(regiao + '.xlsx')
                except:
                    pass
                    
        janela.close()
        
    while ligado == True:
        janela_pesquisa()
        indice = 0
        lin = 1
        number = 1
        number_mes = number
        number = str(number)
        number = "0" + number
        itens_separados = ', '.join(str(item) for item in linhas[indice])
        lista_dos_valores = itens_separados.split(', ')
        lista_geral = []
        regiao = selecionar_regiao()
        resultado = "\n".join(busca())